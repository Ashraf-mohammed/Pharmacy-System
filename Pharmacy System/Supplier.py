import pandas as pd
from openpyxl import load_workbook

def get_id():
    res = ""
    data = pd.read_excel("supplier.xlsx")
    
    for id in data['supplier_id']:
        res = res + str(id) +' , '
    return res
def show_info():
    print('--------- Supplier Form ----------')
    print('----------------------------------')
    print("|Enter 1 to search supplier      |")
    print('----------------------------------')
    print('|Enter 2 to create new supplier  |')
    print('----------------------------------')
    print('|Enter 3 to update supplier info |')
    print('----------------------------------') 
    print('|Enter 4 to go back to main menu |')
    print('----------------------------------')
    
def add_supplier():
    path_file = "supplier.xlsx"
    show_info()
    sup_name = input("Enter New Supplier's Name : ")
    sup_id = input("Enter New Supplier's ID : ")
    sup_city = input("Enter New Supplier's City : ")
    phone = int(input("Enter New Supplier's Contact Number : "))
    sup_email = input("Enter New Supplier's Email : ")
    
    new_row = {'supplier_id':[sup_id],
               'supplier_name':[sup_name],
               'city':[sup_city],
               'phone_number':[phone],
               'email':[sup_email]}
    
    new_df = pd.DataFrame(new_row)
    book = load_workbook(path_file)
    sheet = book.active
    start_row = sheet.max_row + 1
    
    for index, row in new_df.iterrows():
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=start_row + index, column=col_num, value=value)

    # Save the workbook
    book.save(path_file)

    print("New supplier has been added sucessfully.")
    
    
def search_supplier():
    path_file = "supplier.xlsx"
    df = pd.read_excel(path_file)
    sup_id = input("enter supplier ID : ")
    df['supplier_id'] = df['supplier_id'].astype(str).str.strip()
    cnt = 0
    supplier_found = False
    for raw in df['supplier_id']:
        row_index = df[df['supplier_id'] == sup_id].index
        if not row_index.empty :
            row = df.iloc[row_index[0]]
            print("Supplier ID found:", row.to_dict())
            supplier_found = True
            break
        cnt = cnt +1
    
    if not supplier_found:
        print("Supplier ID not found.")
        
def update_supplier():
    path_file = "supplier.xlsx"
    df = pd.read_excel(path_file)
    
    sup_id = input("Enter supplier ID: ")

    print('---------------------------------------------')
    print('| 1. To update name                         |')
    print('---------------------------------------------')
    print('| 2. To update email                   |')
    print('---------------------------------------------')
    print('| 3. To update phone                         |')
    print('---------------------------------------------')
    print('| 4. To update city                  |')
    print('---------------------------------------------')
    choice = int(input('Enter your choice: '))

    # Ensure the column name is correct
    sup_name_column = 'supplier_name'  # Replace this with the correct column name after checking
    sup_city_column = 'city'  # Replace with the correct column name
    sup_email_column = 'email'  # Replace with the correct column name
    sup_phone_column = 'phone_number'  # Replace with the correct column name

    df['supplier_id'] = df['supplier_id'].astype(str).str.strip()
    if sup_id in df['supplier_id'].values:
        if choice == 1:
            new_value = input("Enter the new supplier name: ")
            df.loc[df['supplier_id'] == sup_id, sup_name_column] = new_value
        elif choice == 2:
            new_value = input("Enter the new email : ")
            df.loc[df['supplier_id'] == sup_id, sup_email_column] = new_value
        elif choice == 3:
            new_value = input("Enter the new phone: ")
            df.loc[df['supplier_id'] == sup_id, sup_phone_column] = new_value
        elif choice == 4:
            new_value = input("Enter the new city: ")
            df.loc[df['supplier_id'] == sup_id, sup_city_column] = new_value
        else:
            print("Invalid choice. Please try again.")
        
        # Save the updated DataFrame back to the Excel file
        df.to_excel(path_file, index=False)
        print("Value updated successfully!")
    else:
        print("supplier name not found in the Excel file.")
 