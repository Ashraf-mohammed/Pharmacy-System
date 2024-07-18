import pandas as pd
from openpyxl import load_workbook
import Supplier
def show_menu():
    print('##### Medicine Menu ####')
    print()
    print('---------------------------------------------')
    print("|Enter 1 to add medicine                    |")
    print('---------------------------------------------')
    print('|Enter 2 to search medicine                 |')
    print('---------------------------------------------')
    print('|Enter 3 to update medicine info            |')
    print('---------------------------------------------') 
    print('|Enter 4 for medicines to be purchased list |')
    print('---------------------------------------------')
    print('|Enter 5 to go back to Main Menu            |')
    print('---------------------------------------------')
def add_medicine():
    id = Supplier.get_id()  # this get id of supplier to suggest user to enter inside medicine
    path_file = "medicine.xlsx"
    med_id = input("enter ID :")
    med_name = input("add medicine name :")
    price = input("medicine price :")
    unit = input("unit of medicine :")
    sup_id = input(f"choose one of this supplier id : {id} --------> ")
    
    new_row = {'medicine_id':[med_id],
               'mediciation_name':[med_name],
               'price':[price],
               'unit':[unit],
               'sup_id':[sup_id]}
    
    new_df = pd.DataFrame(new_row)
    book = load_workbook(path_file)
    sheet = book.active
    start_row = sheet.max_row + 1
    
    for index, row in new_df.iterrows():
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=start_row + index, column=col_num, value=value)

    # Save the workbook
    book.save(path_file)

    print("New medicine has been added sucessfully.")
    
def search_medicine():
    path_file = "medicine.xlsx"
    df = pd.read_excel(path_file)
    cnt = 0
    med_name = input("enter medicine name : ")
    rows_with_value = df[df.isin([med_name]).any(axis=1)]
    print("\n",rows_with_value,"\n")
    

def update_medicine():
    path_file = "medicine.xlsx"
    df = pd.read_excel(path_file)
    
    med_name = input("Enter medicine name: ")

    print('---------------------------------------------')
    print('| 1. To update name                         |')
    print('---------------------------------------------')
    print('| 2. To update Cost price                   |')
    print('---------------------------------------------')
    print('| 3. To update unit                         |')
    print('---------------------------------------------')
    print('| 4. To update supplier ID                  |')
    print('---------------------------------------------')
    choice = int(input('Enter your choice: '))

    # Ensure the column name is correct
    med_name_column = 'medication_name'  # Replace this with the correct column name after checking
    cost_price_column = 'price'  # Replace with the correct column name
    unit_column = 'unit'  # Replace with the correct column name
    supplier_id_column = 'sup_id'  # Replace with the correct column name

    if med_name in df[med_name_column].values:
        if choice == 1:
            new_value = input("Enter the new medicine name: ")
            df.loc[df[med_name_column] == med_name, med_name_column] = new_value
        elif choice == 2:
            new_value = input("Enter the new Cost price: ")
            df.loc[df[med_name_column] == med_name, cost_price_column] = new_value
        elif choice == 3:
            new_value = input("Enter the new unit: ")
            df.loc[df[med_name_column] == med_name, unit_column] = new_value
        elif choice == 4:
            new_value = input("Enter the new supplier ID: ")
            df.loc[df[med_name_column] == med_name, supplier_id_column] = new_value
        else:
            print("Invalid choice. Please try again.")
        
        # Save the updated DataFrame back to the Excel file
        df.to_excel(path_file, index=False)
        print("Value updated successfully!")
    else:
        print("Medicine name not found in the Excel file.")
    


def med_info():
    path_file = "medicine.xlsx"
    df = pd.read_excel(path_file)
    
    print(df)
#

    