import pandas as pd
from openpyxl import load_workbook
 
data = pd.read_excel("supplier.xlsx")
print('#' * 20)
print(data)

# New supplier data
new_data = {'supplier_id': [12],
                 'supplier_name': ['ali'],  # Capitalized for consistency
                 'email': ['ali@gmail.com'],
                 'phone_number': ['14578'],
                 'city': ['cairo']}

new_df = pd.DataFrame(new_data)

# Load existing workbook
file_path = 'supplier.xlsx'
book = load_workbook(file_path)

# Select the active sheet (or specify the sheet name)
sheet = book.active

# Find the next row to write the new data
start_row = sheet.max_row + 1
# Append new data to the existing sheet
for index, row in new_df.iterrows():
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=start_row + index, column=col_num, value=value)

# Save the workbook
book.save(file_path)

print("New data has been appended to the existing Excel file.")