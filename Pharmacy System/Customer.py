import pandas as pd
from openpyxl import load_workbook


def show_info():
    
    print('--------- Customer Form ----------')
    print('----------------------------------')
    print("|Enter 1 to search customer      |")
    print('----------------------------------')
    print('|Enter 2 to create new customer  |')
    print('----------------------------------')
    print('|Enter 3 to update customer info |')
    print('----------------------------------') 
    print('|Enter 4 to go back to main menu |')
    print('----------------------------------')
    
    
def add_customer():
    path = ("customer_menu.xlsx")
    customer_name=input('Enter the name of the customer : ')
    customer_id=input('Enter the ID of the customer : ')
    customer_phone=input('Enter the phone number of the customer : ')
    customer_address=input('Enter the address : ')
    
    new_raw = {'cust_id' : [customer_id],
               'cus_name' : [customer_name],
               'phone'  :[customer_phone],
               'city' : [customer_address]}
    
    new_df = pd.DataFrame(new_raw)
    
    book = load_workbook(path)
    sheet = book.active
    start_raw = sheet.max_row + 1
    for index, row in new_df.iterrows():
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=start_raw + index, column=col_num, value=value)

    # Save the workbook
    book.save(path)

    print("New customer has been added sucessfully.")
    
def search_customer():
    path = ("customer_menu.xlsx")
    data = pd.read_excel(path)
    cus_id = input("enter customer ID : ")
    data['cus_id'] = data['cus_id'].astype(str).str.strip()
    res_raw = data[data.isin([cus_id]).any(axis = 1)]
    print('\n',res_raw,'\n')
    
    
def update_customer():
    path_file = "customer_menu.xlsx"
    df = pd.read_excel(path_file)
    
    cus_id = input("Enter customer ID: ")

    print('---------------------------------------------')
    print('| 1. To update name                         |')
    print('---------------------------------------------')
    print('| 2. To update city                         |')
    print('---------------------------------------------')
    print('| 3. To update phone                         |')
    print('---------------------------------------------')
    choice = int(input('Enter your choice: '))

    # Ensure the column name is correct
    cus_name_column = 'cust_name'  # Replace this with the correct column name after checking
    cus_city_column = 'city'  # Replace with the correct column name
    cus_phone_column = 'phone_number'  # Replace with the correct column name

    df['cus_id'] = df['cus_id'].astype(str).str.strip()
    if cus_id in df['cus_id'].values:
        if choice == 1:
            new_value = input("Enter the new customer name: ")
            df.loc[df['cus_id'] == cus_id, cus_name_column] = new_value
        elif choice == 2:
            new_value = input("Enter the new email : ")
            df.loc[df['cus_id'] == cus_id, cus_city_column] = new_value
        elif choice == 3:
            new_value = input("Enter the new phone: ")
            df.loc[df['cus_id'] == cus_id, cus_phone_column] = new_value
        else:
            print("Invalid choice. Please try again.")
        
        # Save the updated DataFrame back to the Excel file
        df.to_excel(path_file, index=False)
        print("Value updated successfully!")
    else:
        print("customer name not found in the Excel file.")
    